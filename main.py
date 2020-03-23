from tkinter import *
from openpyxl import load_workbook, Workbook
from googletrans import Translator

translator = Translator(service_urls=[
      'translate.google.com',
      'translate.google.co.kr',
    ])

def fix_tags(text):
    text = text.replace('</ ', '</')
    text = text.replace(' <', '<')
    text = text.replace('> ', '>')
    return text

def translate(text):
    return fix_tags(translator.translate([str(text)], src="ru", dest="uk")[0].text)

def translate_clicked():
    filename = workbook_field.get(1.0, END).replace('\n','')
    sheet_to_read = worksheet_field.get(1.0, END).replace('\n','')
    columns_to_read = columns_field.get(1.0, END).replace('\n','').split(' ')

    wb = load_workbook(filename)
    ws = wb[sheet_to_read]
    wb_ukr = Workbook()
    wb_ukr.create_sheet('ukr')
    ws_ukr = wb_ukr['ukr']
    columns = [ws[column] for column in columns_to_read]

    for col_num in range(0, len(columns)):
        for row_num in range(0, len(columns[col_num])):
            text = columns[col_num][row_num].value
            if not text:
                ws_ukr.cell(row=row_num+1,column=col_num+1).value = ''
                continue
            try:
                ws_ukr.cell(row=row_num+1,column=col_num+1).value = translate(text)
            except:
                ws_ukr.cell(row=row_num+1,column=col_num+1).value = text

    wb_ukr.save("translated_{}".format(filename))

window = Tk()
window.title('XL Translate')
window.geometry('400x200')
window.resizable(0, 0)

workbook_label = Label(window, text="Workbook")
workbook_label.pack()

workbook_field = Text(window, width=40, height = 1)
workbook_field.pack()

worksheet_label = Label(window, text="Worksheet")
worksheet_label.pack()

worksheet_field = Text(window, width=40, height = 1)
worksheet_field.pack()

columns_label = Label(window, text="Columns")
columns_label.pack()

columns_field = Text(window, width=40, height = 1)
columns_field.pack()

translate_button = Button(window, text="TRANSLATE", bg = 'blue', fg ='white', command=translate_clicked)
translate_button.pack()

window.mainloop()
